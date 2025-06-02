import csv
import random
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# --- Membaca Dataset dari file CSV ---
def load_dataset(filename, split_ratio):
    with open(filename, 'r') as file:
        lines = csv.reader(file)
        dataset = []
        next(lines)  # Lewati header
        for row in lines:
            if len(row) < 6:
                continue
            try:
                features = [float(x) for x in row[1:5]]  # Kolom fitur
                label = row[5]  # Kolom label
                dataset.append(features + [label])
            except ValueError:
                continue
        print(f"[DEBUG] Total data terbaca: {len(dataset)}")
        random.shuffle(dataset)
        split_index = int(split_ratio * len(dataset))
        return dataset[:split_index], dataset[split_index:]

# --- Hitung Jarak Euclidean ---
def euclidean_distance(data1, data2):
    return math.sqrt(sum((data1[i] - data2[i]) ** 2 for i in range(len(data1) - 1)))

# --- Cari K Tetangga Terdekat ---
def get_neighbors(training_set, test_instance, k):
    distances = [(train, euclidean_distance(test_instance, train)) for train in training_set]
    distances.sort(key=lambda x: x[1])
    return [distances[i][0] for i in range(k)]

# --- Voting Mayoritas ---
def predict_classification(neighbors):
    votes = {}
    for neighbor in neighbors:
        label = neighbor[-1]
        votes[label] = votes.get(label, 0) + 1
    return sorted(votes.items(), key=lambda x: x[1], reverse=True)[0][0]

# --- Hitung Metrik seluruh model ---
def calculate_accuracy(test_set, predictions):
    correct = sum(1 for i in range(len(test_set)) if test_set[i][-1] == predictions[i])
    return correct / len(test_set) * 100.0

def calculate_recall(test_set, predictions):
    actual = [row[-1] for row in test_set]
    labels = sorted(list(set(actual + predictions)))
    recall_dict = {}
    for label in labels:
        true_pos = sum(1 for a, p in zip(actual, predictions) if a == label and p == label)
        false_neg = sum(1 for a, p in zip(actual, predictions) if a == label and p != label)
        denominator = true_pos + false_neg
        recall = (true_pos / denominator * 100.0) if denominator > 0 else 0.0
        recall_dict[label] = f"{recall:.2f}%"
    return recall_dict

def calculate_true_negative_rate(test_set, predictions):
    actual = [row[-1] for row in test_set]
    labels = sorted(list(set(actual + predictions)))
    true_negatives = {}
    for label in labels:
        true_neg = sum(1 for a, p in zip(actual, predictions) if a != label and p != label)
        total_negatives = sum(1 for a in actual if a != label)
        true_negative_rate = (true_neg / total_negatives * 100.0) if total_negatives > 0 else 0.0
        true_negatives[label] = f"{true_negative_rate:.2f}%"
    return true_negatives

def caluclate_f1_score(test_set, predictions):
    precision = calculate_precision(test_set, predictions)
    recall = calculate_recall(test_set, predictions)
    f1_scores = {}
    for label in precision:
        p = float(precision[label].strip('%')) / 100.0
        r = float(recall[label].strip('%')) / 100.0
        if p + r > 0:
            f1_scores[label] = f"{2 * (p * r) / (p + r) * 100:.2f}%"
        else:
            f1_scores[label] = "0.00%"
    return f1_scores

def calculate_precision(test_set, predictions):
    actual = [row[-1] for row in test_set]
    labels = sorted(list(set(actual + predictions)))
    precision_dict = {}
    for label in labels:
        true_pos = sum(1 for a, p in zip(actual, predictions) if a == label and p == label)
        false_pos = sum(1 for a, p in zip(actual, predictions) if a != label and p == label)
        denominator = true_pos + false_pos
        precision = (true_pos / denominator * 100.0) if denominator > 0 else 0.0
        precision_dict[label] = f"{precision:.2f}%"
    return precision_dict

# --- Buat Confusion Matrix Manual ---
def create_confusion_matrix(actual, predicted):
    labels = sorted(set(actual + predicted))
    matrix = {label: {l: 0 for l in labels} for label in labels}
    for a, p in zip(actual, predicted):
        matrix[a][p] += 1
    return labels, matrix

# --- Simpan Semua Hasil ke Excel ---
def save_all_to_excel(test_set, predictions):
    actual = [row[-1] for row in test_set]
    predicted = predictions
    labels, cmatrix = create_confusion_matrix(actual, predicted)

    wb = Workbook()

    # Sheet 1 - Hasil Prediksi
    ws1 = wb.active
    ws1.title = "Hasil Prediksi"
    headers1 = ['SepalLengthCm', 'SepalWidthCm', 'PetalLengthCm', 'PetalWidthCm', 'Actual', 'Predicted']
    ws1.append(headers1)

    fill_benar = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # hijau
    fill_salah = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # merah

    for i in range(len(test_set)):
        features = test_set[i][:-1]
        a = actual[i]
        p = predicted[i]
        row = features + [a, p]
        ws1.append(row)
        for col in range(1, 7):
            ws1.cell(row=i+2, column=col).fill = fill_benar if a == p else fill_salah

    # Sheet 2 - Confusion Matrix
    ws2 = wb.create_sheet(title="Confusion Matrix")
    ws2.append([""] + labels)
    for label in labels:
        row = [label] + [cmatrix[label][l] for l in labels]
        ws2.append(row)

    # Sheet 3 - Statistik Per Kelas
    ws3 = wb.create_sheet(title="Statistik Per Kelas")
    ws3.append(["Kelas", "Jumlah Aktual", "True Positive", "Jumlah Diprediksi", "Akurasi (%)"])
    for label in labels:
        total_aktual = sum(cmatrix[label].values())
        true_positive = cmatrix[label][label]
        total_prediksi = sum(cmatrix[l][label] for l in labels)
        akurasi = (true_positive / total_aktual) * 100 if total_aktual > 0 else 0
        ws3.append([label, total_aktual, true_positive, total_prediksi, round(akurasi, 2)])

    # Simpan file dengan nama tetap
    filename = 'hasil_prediksi_iris.xlsx'
    wb.save(filename)
    print(f"\n[INFO] Hasil lengkap disimpan di: {filename}")

def input_values():
    try:
        sepal_length = float(input("Masukkan Sepal Length (cm): "))
        sepal_width = float(input("Masukkan Sepal Width (cm): "))
        petal_length = float(input("Masukkan Petal Length (cm): "))
        petal_width = float(input("Masukkan Petal Width (cm): "))
        return [sepal_length, sepal_width, petal_length, petal_width]
    except ValueError:
        print("Input tidak valid. Silakan masukkan angka yang benar.")
        return input_values()

def predict_user_input(training_set, k):
    user_input = input_values()
    neighbors = get_neighbors(training_set, user_input + [''], k)
    result = predict_classification(neighbors)
    print(f'Kelas terprediksi untuk {user_input} adalah: {result}')
    return result

# --- Fungsi Utama ---
def main():
    filename = 'Iris.csv'
    split_ratio = 0.7 # 70% data untuk training, 30% untuk testing
    k = 3

    training_set, testing_set = load_dataset(filename, split_ratio)
    print(f'Training: {len(training_set)} data')
    print(f'Testing : {len(testing_set)} data\n')

    predictions = []
    for test in testing_set:
        neighbors = get_neighbors(training_set, test, k)
        result = predict_classification(neighbors)
        predictions.append(result)
        print(f'Predicted={result} \tActual={test[-1]}')

    training_predictions = []
    for train in training_set:
        neighbors = get_neighbors(training_set, train, k)
        result = predict_classification(neighbors)
        training_predictions.append(result)

    accuracy = calculate_accuracy(testing_set, predictions) # Testing accuracy
    training_accuracy = calculate_accuracy(training_set, training_predictions)

    testing_recall = calculate_recall(testing_set, predictions)
    training_recall = calculate_recall(training_set, training_predictions)

    test_f1_scores = caluclate_f1_score(testing_set, predictions)
    train_f1_scores = caluclate_f1_score(training_set, training_predictions)

    test_true_negatives = calculate_true_negative_rate(testing_set, predictions)
    train_true_negatives = calculate_true_negative_rate(training_set, training_predictions)

    training_precision = calculate_precision(training_set, training_predictions)
    testing_precision = calculate_precision(testing_set, predictions)

    print(f'\n=== METRIK MODEL ===')
    print(f'Testing Accuracy   : {accuracy:.2f}%')
    print(f'Training Accuracy  : {training_accuracy:.2f}%')

    print('\nRecall per kelas')
    print('Testing Recall   :', ', '.join([f"{k}: {v}" for k, v in testing_recall.items()]))
    print('Training Recall  :', ', '.join([f"{k}: {v}" for k, v in training_recall.items()]))

    print('\nPrecision per kelas:')
    print('Testing Precision :', ', '.join([f"{k}: {v}" for k, v in testing_precision.items()]))
    print('Training Precision:', ', '.join([f"{k}: {v}" for k, v in training_precision.items()]))

    print('\nSpecificity per kelas:')
    print('Testing Specificity :', ', '.join([f"{k}: {v}" for k, v in test_true_negatives.items()]))
    print('Training Specificity:', ', '.join([f"{k}: {v}" for k, v in train_true_negatives.items()]))

    print('\nF1 Score per kelas:')
    print('Testing F1 Score   :', ', '.join([f"{k}: {v}" for k, v in test_f1_scores.items()]))
    print('Training F1 Score  :', ', '.join([f"{k}: {v}" for k, v in train_f1_scores.items()]))

    save_all_to_excel(testing_set, predictions)

    print("Memprediksi input pengguna...")
    user_prediction = predict_user_input(training_set, k)
    print(f'Prediksi untuk input pengguna: {user_prediction}')

if __name__ == '__main__':
    main()
